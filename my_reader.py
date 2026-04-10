import openpyxl
from DataDriver.AbstractReaderClass import AbstractReaderClass
from DataDriver.ReaderConfig import TestCaseData

class my_reader(AbstractReaderClass):
    
    def get_data_from_source(self):
        sheet_name = self.kwargs.get('sheet_name', 'Sheet1')
        file_path = self.file
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        
        # 1. อ่านหัว Column (Row 1)
        raw_headers = [str(cell.value).strip() if cell.value else None for cell in sheet[1]]
        
        test_data = []
        
        # 2. อ่านข้อมูลตั้งแต่แถวที่ 2
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_dict = {}
            tags = []
            doc = ""
            case_name = None
            
            # 3. วนลูปเช็คแต่ละ Cell ในแถวเพื่อแยกประเภทข้อมูล
            for i, column_name in enumerate(raw_headers):
                if not column_name: continue
                
                value = row[i] if i < len(row) else ""
                
                # ตรวจสอบหัวข้อพิเศษตามมาตรฐาน DataDriver
                if column_name.lower().strip() == "*** test cases ***" or column_name.lower().strip() == "*** tasks ***":
                    case_name = str(value) if value else None
                elif column_name.lower().strip() == "[tags]":
                    if value:
                        # รองรับการแยก tag ด้วย comma หรือ semicolon
                        tags = [t.strip() for t in str(value).replace(';', ',').split(',')]
                elif column_name.lower().strip() == "[documentation]":
                    doc = str(value) if value else ""
                else:
                    # ถ้าไม่ใช่หัวข้อพิเศษ ให้มองเป็นตัวแปร (${VarName})
                    clean_key = f"${{{column_name}}}" if not column_name.startswith("${") else column_name
                    row_dict[clean_key] = value
            
            # 4. สร้าง TestCaseData พร้อมส่ง Tags และ Documentation กลับไป
            test_data.append(
                TestCaseData(
                    test_case_name=case_name, 
                    arguments=row_dict,
                    tags=tags,
                    documentation=doc
                )
            )
            
        return test_data